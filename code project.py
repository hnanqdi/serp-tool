import streamlit as st
import pandas as pd
import numpy as np
import json, os, io, re, tempfile
from urllib.parse import urlparse
from sklearn.cluster import KMeans
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import requests

st.set_page_config(page_title="ابزار پیشنهاد سایت رپورتاژ", layout="wide")

# ---------- Custom Theme CSS ----------
st.markdown("""
<style>
body {
    background-color: #F4E9D7;
    color: #000000;
}
div[data-testid="stSidebar"] {
    background-color: #B8C4A9;
}
div.stButton > button:first-child {
    background-color: #D97D55;
    color: white;
    border-radius: 10px;
    height: 3em;
    width: 100%;
    font-weight: bold;
}
div.stDownloadButton > button {
    background-color: #6FA4AF;
    color: white;
    border-radius: 10px;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

st.title("ابزار پیشنهاد سایت رپورتاژ")

# ---------- Helper functions ----------
def safe_filename(name: str):
    return re.sub(r'[<>:"/\\|?*]', '_', name)

def read_keywords_file(uploaded):
    if uploaded is None: return []
    s = uploaded.getvalue().decode('utf-8', errors='ignore')
    return [line.strip() for line in s.splitlines() if line.strip()]

def read_urls_df(uploaded):
    if uploaded is None:
        raise ValueError("فایل urls_list.xlsx آپلود نشده است.")
    return pd.read_excel(uploaded)

def extract_domains_from_urls_df(df):
    col = next((c for c in df.columns if 'url' in str(c).lower() or 'link' in str(c).lower() or 'domain' in str(c).lower()), df.columns[0])
    urls = df[col].dropna().astype(str).tolist()
    domains = set()
    for u in urls:
        u2 = u if re.match(r'https?://', u) else 'http://' + u
        try: domains.add(urlparse(u2).netloc.lower().lstrip('www.'))
        except: continue
    return domains

# ---------- Fetch SERP data with batching ----------
def fetch_serp_data_batch(keywords, api_key, output_folder, batch_size=50):
    if not os.path.exists(output_folder): os.makedirs(output_folder)
    url = 'https://google.serper.dev/search'
    headers = {
        'X-API-KEY': api_key,
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
    }
    for i in range(0, len(keywords), batch_size):
        batch = keywords[i:i+batch_size]
        payload = json.dumps([{"q": kw, "gl": "ir", "num": 50, "autocorrect": True} for kw in batch])
        try:
            response = requests.post(url, headers=headers, data=payload)
            if response.status_code == 200:
                serp_responses = response.json()
                for j, keyword in enumerate(batch):
                    data = serp_responses[j]
                    with open(os.path.join(output_folder, safe_filename(keyword)+'.json'), 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False)
            else:
                st.warning(f"خطا در دریافت داده برای batch {i}: {response.status_code}")
        except Exception as e:
            st.warning(f"خطا در درخواست Serper برای batch {i}: {e}")

# ---------- Process JSON ----------
def process_json_folder(data_folder, domains_to_check):
    results = []
    for fn in os.listdir(data_folder):
        if not fn.lower().endswith('.json'): continue
        path = os.path.join(data_folder, fn)
        try: data = json.load(open(path,'r',encoding='utf-8'))
        except: continue
        query = data.get('searchParameters', {}).get('q', '')
        organic = data.get('organic', [])
        for res in organic:
            link = res.get('link','') or res.get('url','')
            if not link: continue
            domain = urlparse(link if re.match(r'https?://', link) else 'http://' + link).netloc.lower().lstrip('www.')
            if domain in domains_to_check:
                position = res.get('position', None)
                title = res.get('title', '')
                results.append([domain, query, position, title, link])
    df = pd.DataFrame(results, columns=['Domain','Query','Position','Title','Link'])
    df['Position'] = pd.to_numeric(df['Position'], errors='coerce')
    return df

# ---------- Clustering & Score ----------
def cluster_domains_from_results(df_results):
    if df_results.empty: return pd.DataFrame(columns=['Domain','Position','AppearanceCount','Cluster','ClusterLabel','Score'])
    domain_grouped = df_results.groupby('Domain').agg({'Position':'mean','Query':'count'}).reset_index().rename(columns={'Query':'AppearanceCount'})
    X_numeric = domain_grouped[['Position','AppearanceCount']].fillna(999).values
    kmeans = KMeans(n_clusters=2, random_state=42)
    domain_grouped['Cluster'] = kmeans.fit_predict(X_numeric)
    cluster_summary = domain_grouped.groupby('Cluster')[['Position','AppearanceCount']].mean()
    order = cluster_summary.sort_values('Position').index.tolist()
    mapping = {cl:'قوی' if i==0 else 'ضعیف' for i, cl in enumerate(order)}
    domain_grouped['ClusterLabel'] = domain_grouped['Cluster'].map(mapping)
    domain_grouped['Score'] = domain_grouped.apply(lambda row: row['AppearanceCount']/row['Position'] if row['Position']>0 else 999, axis=1)
    return domain_grouped.sort_values('Score', ascending=False).reset_index(drop=True)

# ---------- Excel Builder ----------
def build_excel_bytes(df_final, domain_grouped):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Results', index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb['Results']
    for cell in ws[1]: cell.font = Font(bold=True)
    cluster_col = next((idx+1 for idx, cell in enumerate(ws[1]) if str(cell.value)=='ClusterLabel'), None)
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    if cluster_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            val = row[cluster_col-1].value
            fill = green if val=='قوی' else (red if val=='ضعیف' else None)
            if fill: [setattr(c,'fill',fill) for c in row]
    ws.freeze_panes='A2'
    for i, col in enumerate(ws.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(cell.value)) if cell.value else 0 for cell in col)+2,80)
    heat_ws = wb.create_sheet(title='Heatmap')
    queries = df_final['Query'].unique().tolist() if not df_final.empty else []
    for col_idx, q in enumerate(['Domain']+queries, start=1):
        heat_ws.cell(row=1, column=col_idx, value=q)
    for r_idx, domain in enumerate(domain_grouped['Domain'].tolist(), start=2):
        heat_ws.cell(row=r_idx, column=1, value=domain)
        for c_idx, query in enumerate(queries, start=2):
            pos_series = df_final[(df_final['Domain']==domain)&(df_final['Query']==query)]['Position']
            if not pos_series.empty:
                v = pos_series.values[0]
                cell = heat_ws.cell(row=r_idx,column=c_idx,value=v)
                if v<=3: cell.fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                elif v<=10: cell.fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else: cell.fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    bc_ws = wb.create_sheet(title='BarChart')
    bc_ws.cell(row=1,column=1,value='Domain')
    bc_ws.cell(row=1,column=2,value='Score')
    for idx,row in enumerate(domain_grouped.itertuples(), start=2):
        bc_ws.cell(row=idx,column=1,value=row.Domain)
        bc_ws.cell(row=idx,column=2,value=row.Score)
    chart = BarChart()
    chart.title = 'دامنه‌ها از قوی‌ترین به ضعیف‌ترین'
    data = Reference(bc_ws,min_col=2,min_row=2,max_row=1+len(domain_grouped))
    cats = Reference(bc_ws,min_col=1,min_row=2,max_row=1+len(domain_grouped))
    chart.add_data(data,titles_from_data=False)
    chart.set_categories(cats)
    bc_ws.add_chart(chart,'D2')
    out = io.BytesIO(); wb.save(out); out.seek(0); return out

# ---------- UI ----------
st.sidebar.header('ورودی‌ها')
st.sidebar.markdown('**آپلود فایل کلمات کلیدی (.txt)**')
uploaded_keywords = st.sidebar.file_uploader('', type=['txt'])
st.sidebar.markdown('**آپلود فایل لیست دامنه (.xlsx)**')
uploaded_urls = st.sidebar.file_uploader('', type=['xlsx'])
api_key = st.sidebar.text_input('API Key Serper.dev')
run_button = st.button('اجرای پردازش')

if run_button:
    if uploaded_urls is None or uploaded_keywords is None or not api_key:
        st.error('لطفاً همه ورودی‌ها شامل keywords.txt، urls_list.xlsx و API key را وارد کنید.')
    else:
        with st.spinner('در حال گرفتن داده‌ها و پردازش ...'):
            tmpdir = tempfile.mkdtemp(prefix='serp_streamlit_')
            keywords = read_keywords_file(uploaded_keywords)
            urls_df = read_urls_df(uploaded_urls)
            domains = extract_domains_from_urls_df(urls_df)
            fetch_serp_data_batch(keywords, api_key, tmpdir)  # درخواست‌ها به‌صورت batch
            df_results = process_json_folder(tmpdir, domains)
            domain_grouped = cluster_domains_from_results(df_results)
            df_final = df_results.merge(domain_grouped[['Domain','ClusterLabel']], on='Domain', how='left') if not df_results.empty else pd.DataFrame(columns=['Domain','Query','Position','Title','Link','ClusterLabel'])
            excel_bytes = build_excel_bytes(df_final, domain_grouped)
            st.success('پردازش انجام شد')
            st.dataframe(domain_grouped.head(50))
            st.download_button('دانلود فایل اکسل', data=excel_bytes.getvalue(), file_name='serp_reportage_tool.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
